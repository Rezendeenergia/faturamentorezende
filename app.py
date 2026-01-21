from flask import Flask, render_template, request, jsonify, send_file
import os
import sqlite3
from werkzeug.utils import secure_filename
from ocr_extractor import NFExtractor
from calculadora_retencoes import CalculadoraRetencoes
from database import Database
import pandas as pd
from datetime import datetime
import tempfile

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Cria diretório de uploads se não existir
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf'}

# Inicializa banco de dados
db = Database()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """Recebe PDF e extrai dados via OCR"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Arquivo vazio'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Apenas arquivos PDF são permitidos'}), 400

    try:
        # Salva arquivo temporariamente
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Extrai dados via OCR
        extractor = NFExtractor(filepath)
        dados_extraidos = extractor.extract()

        # Calcula retenções e valores usando a calculadora independente
        calc = CalculadoraRetencoes()
        retencoes = calc.calcular_retencoes(
            dados_extraidos['tipo'],
            dados_extraidos['valor_bruto'],
            pis_cofins_retido=False
        )
        valor_nominal = calc.calcular_valor_nominal(dados_extraidos['valor_bruto'], retencoes)

        # Adiciona cálculos aos dados extraídos
        dados_extraidos['retencoes'] = retencoes
        dados_extraidos['valor_nominal_calculado'] = round(valor_nominal, 2)

        # Remove arquivo temporário
        os.remove(filepath)

        return jsonify({
            'success': True,
            'dados': dados_extraidos
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500


@app.route('/salvar', methods=['POST'])
def salvar_nota():
    """Salva nota fiscal no banco de dados"""
    try:
        dados = request.json

        # Prepara dados para o banco (formato YYYY-MM-DD)
        dados_db = dados.copy()
        data_parts = dados['data_emissao'].split('/')
        dados_db['data_emissao'] = f"{data_parts[2]}-{data_parts[1]}-{data_parts[0]}"

        # Adiciona valores que vieram do formulário (editáveis)
        dados_db['inss'] = dados.get('inss', 0)
        dados_db['iss'] = dados.get('iss', 0)
        dados_db['retencao_equatorial'] = dados.get('retencao_equatorial', 0)
        dados_db['pis_cofins_csll'] = dados.get('pis_cofins_csll', 0)
        dados_db['pis_cofins_retido'] = dados.get('pis_cofins_retido', False)
        dados_db['valor_nominal_calculado'] = dados.get('valor_nominal_calculado', 0)
        dados_db['valor_nominal_conferencia'] = dados.get('valor_nominal_conferencia', 0)

        # Insere no banco
        db.inserir_nota(dados_db)

        return jsonify({
            'success': True,
            'message': f'Nota fiscal {dados["numero_nf"]} salva com sucesso!'
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao salvar: {str(e)}'}), 500


@app.route('/calcular', methods=['POST'])
def calcular_valores():
    """Calcula valores baseado no tipo e valor bruto"""
    try:
        dados = request.json
        tipo = dados.get('tipo')
        valor_bruto = float(dados.get('valor_bruto', 0))
        pis_cofins_retido = dados.get('pis_cofins_retido', False)

        # Usa calculadora independente
        calc = CalculadoraRetencoes()
        retencoes = calc.calcular_retencoes(tipo, valor_bruto, pis_cofins_retido)
        valor_nominal = calc.calcular_valor_nominal(valor_bruto, retencoes)

        return jsonify({
            'success': True,
            'retencoes': {
                'inss': round(retencoes['inss'], 2),
                'iss': round(retencoes['iss'], 2),
                'retencao_equatorial': round(retencoes['retencao_equatorial'], 2),
                'pis_cofins_csll': round(retencoes['pis_cofins_csll'], 2)
            },
            'valor_nominal': round(valor_nominal, 2)
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao calcular: {str(e)}'}), 500


@app.route('/dashboard')
def dashboard():
    """Página do dashboard de recebimentos"""
    return render_template('dashboard.html')


@app.route('/extrato')
def extrato():
    """Página de gerenciamento de extrato"""
    return render_template('extrato.html')


@app.route('/api/dashboard-data')
def dashboard_data():
    """Retorna dados para o dashboard"""
    try:
        data = db.dashboard_recebimentos()
        pendentes = db.listar_pendentes()
        analise = db.analise_financeira()

        return jsonify({
            'success': True,
            'dashboard': data,
            'pendentes': pendentes,
            'analise_financeira': analise
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notas-pendentes')
def notas_pendentes():
    """Lista notas pendentes de recebimento"""
    try:
        notas = db.listar_pendentes()
        return jsonify({
            'success': True,
            'notas': notas
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/todas-notas')
def todas_notas():
    """Lista todas as notas fiscais"""
    try:
        notas = db.listar_todas_notas()
        return jsonify({
            'success': True,
            'notas': notas
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/extrato')
def listar_extrato():
    """Lista lançamentos do extrato com filtro opcional"""
    try:
        # Pega filtro da query string
        filtro = request.args.get('filtro_adiantamento')

        if filtro == 'adiantados':
            extrato = db.listar_extrato(filtro_adiantamento=True)
        elif filtro == 'normais':
            extrato = db.listar_extrato(filtro_adiantamento=False)
        else:
            extrato = db.listar_extrato()

        return jsonify({
            'success': True,
            'extrato': extrato
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/registrar-recebimento', methods=['POST'])
def registrar_recebimento():
    """Registra um novo recebimento no extrato"""
    try:
        dados = request.json

        # Converte data para formato YYYY-MM-DD
        data_parts = dados['data_recebimento'].split('/')
        dados['data_recebimento'] = f"{data_parts[2]}-{data_parts[1]}-{data_parts[0]}"

        extrato_id = db.inserir_recebimento(dados)

        return jsonify({
            'success': True,
            'message': 'Recebimento registrado com sucesso!',
            'extrato_id': extrato_id
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao registrar recebimento: {str(e)}'}), 500


@app.route('/importar')
def importar_page():
    """Página de importação"""
    return render_template('importar.html')


@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    """Recebe arquivo Excel e importa dados"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Arquivo vazio'}), 400

    if not file.filename.endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Apenas arquivos Excel (.xlsx, .xlsm) são permitidos'}), 400

    try:
        # Salva arquivo temporariamente
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        # Importa dados
        from importar_planilha import PlanilhaImporter
        importer = PlanilhaImporter(tmp_path)
        resultado = importer.importar_tudo()

        # Remove arquivo temporário
        os.remove(tmp_path)

        return jsonify({
            'success': True,
            'message': 'Planilha importada com sucesso!',
            'dados': {
                'nfs': resultado['nfs'],
                'extrato': resultado['extrato'],
                'dashboard': {
                    'recebido': {
                        'qtd': resultado['dashboard']['recebido']['qtd'],
                        'total': float(resultado['dashboard']['recebido']['total'] or 0)
                    },
                    'a_receber': {
                        'qtd': resultado['dashboard']['a_receber']['qtd'],
                        'total': float(resultado['dashboard']['a_receber']['total'] or 0)
                    },
                    'atrasado': {
                        'qtd': resultado['dashboard']['atrasado']['qtd'],
                        'total': float(resultado['dashboard']['atrasado']['total'] or 0)
                    }
                }
            }
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao importar: {str(e)}'}), 500


@app.route('/adiantar')
def adiantar_page():
    """Página de adiantamento de NFs"""
    return render_template('adiantar_nf.html')


@app.route('/api/adiantar-nota', methods=['POST'])
def adiantar_nota():
    """Registra adiantamento de uma nota fiscal"""
    try:
        dados = request.json

        resultado = db.adiantar_nota(dados['nota_id'], dados)

        return jsonify({
            'success': True,
            'message': 'Adiantamento registrado com sucesso e lançamento criado no extrato!',
            'dados': resultado
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao registrar adiantamento: {str(e)}'}), 500


@app.route('/api/exportar/<tipo_relatorio>')
def exportar_relatorio(tipo_relatorio):
    """Exporta relatório para Excel"""
    try:
        # Obtém dados do banco
        dados = db.exportar_para_excel(tipo_relatorio)

        if not dados:
            return jsonify({'error': 'Nenhum dado para exportar'}), 400

        # Cria DataFrame
        df = pd.DataFrame(dados)

        # Cria arquivo temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            # Salva Excel
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Dados')

                # Formata worksheet
                worksheet = writer.sheets['Dados']

                # Ajusta largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            tmp_path = tmp.name

        # Gera nome do arquivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f'relatorio_{tipo_relatorio}_{timestamp}.xlsx'

        # Envia arquivo
        return send_file(
            tmp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        return jsonify({'error': f'Erro ao exportar: {str(e)}'}), 500


@app.route('/api/exportar-completo')
def exportar_completo():
    """Exporta planilha completa com abas NF'S e Extrato (formato original)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Cria workbook
        wb = Workbook()

        # ============================================================
        # ABA 1: NF'S
        # ============================================================
        ws_nfs = wb.active
        ws_nfs.title = "NF'S"

        # Cabeçalhos (linha 1)
        headers = [
            'Data Emissão',  # A
            'Nº NF',  # B
            'Tipo',  # C
            'Valor Bruto',  # D
            'Localidade',  # E
            'Retenções Federais (INSS)',  # F
            'Alíquota INSS',  # G
            'ISS',  # H
            'Alíquota ISS',  # I
            'Retenção Equatorial',  # J
            'Tomador do Serviço',  # K
            'PIS/COFINS/CSLL',  # L
            'Valor Nominal Conferência',  # M
            'Valor Nominal (Vinci)',  # N
            'Valor Líquido Vinci',  # O
            'Data do adiantamento',  # P
            '% de Adiantamento',  # Q
            'Valor retido Vinci'  # R
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_nfs.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Busca dados do banco
        conn = sqlite3.connect(db.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT 
                data_emissao, numero_nf, tipo, valor_bruto, localidade,
                inss, iss, retencao_equatorial, tomador, pis_cofins_csll,
                valor_nominal_conferencia, valor_nominal_calculado, 
                valor_liquido_vinci, foi_adiantado, data_adiantamento,
                percentual_adiantamento, valor_retido_vinci
            FROM notas_fiscais
            ORDER BY data_emissao, numero_nf
        ''')

        notas = cursor.fetchall()

        # Preenche dados
        for row_idx, nota in enumerate(notas, 2):
            # A - Data Emissão
            if nota['data_emissao']:
                ws_nfs.cell(row_idx, 1, nota['data_emissao'])

            # B - Nº NF
            ws_nfs.cell(row_idx, 2, nota['numero_nf'])

            # C - Tipo
            ws_nfs.cell(row_idx, 3, nota['tipo'])

            # D - Valor Bruto
            ws_nfs.cell(row_idx, 4, nota['valor_bruto'])
            ws_nfs.cell(row_idx, 4).number_format = 'R$ #,##0.00'

            # E - Localidade
            ws_nfs.cell(row_idx, 5, nota['localidade'])

            # F - INSS
            ws_nfs.cell(row_idx, 6, nota['inss'] or 0)
            ws_nfs.cell(row_idx, 6).number_format = 'R$ #,##0.00'

            # G - Alíquota INSS (FÓRMULA)
            ws_nfs.cell(row_idx, 7, f'=F{row_idx}/D{row_idx}')
            ws_nfs.cell(row_idx, 7).number_format = '0.00%'

            # H - ISS
            ws_nfs.cell(row_idx, 8, nota['iss'] or 0)
            ws_nfs.cell(row_idx, 8).number_format = 'R$ #,##0.00'

            # I - Alíquota ISS (FÓRMULA)
            ws_nfs.cell(row_idx, 9, f'=H{row_idx}/D{row_idx}')
            ws_nfs.cell(row_idx, 9).number_format = '0.00%'

            # J - Retenção Equatorial
            ws_nfs.cell(row_idx, 10, nota['retencao_equatorial'] or 0)
            ws_nfs.cell(row_idx, 10).number_format = 'R$ #,##0.00'

            # K - Tomador
            ws_nfs.cell(row_idx, 11, nota['tomador'])

            # L - PIS/COFINS/CSLL
            ws_nfs.cell(row_idx, 12, nota['pis_cofins_csll'] or 0)
            ws_nfs.cell(row_idx, 12).number_format = 'R$ #,##0.00'

            # M - Valor Nominal Conferência
            ws_nfs.cell(row_idx, 13, nota['valor_nominal_conferencia'] or nota['valor_nominal_calculado'] or 0)
            ws_nfs.cell(row_idx, 13).number_format = 'R$ #,##0.00'

            # N - Valor Nominal (Vinci)
            ws_nfs.cell(row_idx, 14, nota['valor_nominal_calculado'] or 0)
            ws_nfs.cell(row_idx, 14).number_format = 'R$ #,##0.00'

            # O - Valor Líquido Vinci
            if nota['valor_liquido_vinci']:
                ws_nfs.cell(row_idx, 15, nota['valor_liquido_vinci'])
                ws_nfs.cell(row_idx, 15).number_format = 'R$ #,##0.00'

            # P - Data do adiantamento
            if nota['data_adiantamento']:
                ws_nfs.cell(row_idx, 16, nota['data_adiantamento'])

            # Q - % de Adiantamento (FÓRMULA)
            ws_nfs.cell(row_idx, 17, f'=IF(O{row_idx}>0,(N{row_idx}-O{row_idx})/N{row_idx},"")')
            ws_nfs.cell(row_idx, 17).number_format = '0.00%'

            # R - Valor retido Vinci (FÓRMULA)
            ws_nfs.cell(row_idx, 18, f'=IF(O{row_idx}>0,N{row_idx}-O{row_idx},"")')
            ws_nfs.cell(row_idx, 18).number_format = 'R$ #,##0.00'

        # Ajusta largura das colunas
        column_widths = {
            'A': 15, 'B': 12, 'C': 20, 'D': 15, 'E': 20,
            'F': 18, 'G': 15, 'H': 12, 'I': 15, 'J': 18,
            'K': 35, 'L': 18, 'M': 22, 'N': 20, 'O': 20,
            'P': 18, 'Q': 18, 'R': 20
        }
        for col, width in column_widths.items():
            ws_nfs.column_dimensions[col].width = width

        # ============================================================
        # ABA 2: EXTRATO
        # ============================================================
        ws_extrato = wb.create_sheet("Extrato")

        # Cabeçalhos
        extrato_headers = [
            'Data',
            'Valor',
            "NF'S",
            'Tipo',
            'Complemento'
        ]

        for col, header in enumerate(extrato_headers, 1):
            cell = ws_extrato.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Busca dados do extrato
        cursor.execute('''
            SELECT 
                data_recebimento, valor_recebido, nfs_referentes,
                tipo_recebimento, complemento
            FROM extrato
            ORDER BY data_recebimento
        ''')

        extratos = cursor.fetchall()

        # Preenche dados
        for row_idx, extrato in enumerate(extratos, 2):
            ws_extrato.cell(row_idx, 1, extrato['data_recebimento'])
            ws_extrato.cell(row_idx, 2, extrato['valor_recebido'])
            ws_extrato.cell(row_idx, 2).number_format = 'R$ #,##0.00'
            ws_extrato.cell(row_idx, 3, extrato['nfs_referentes'])
            ws_extrato.cell(row_idx, 4, extrato['tipo_recebimento'])
            ws_extrato.cell(row_idx, 5, extrato['complemento'] or '')

        # Ajusta largura das colunas
        ws_extrato.column_dimensions['A'].width = 15
        ws_extrato.column_dimensions['B'].width = 18
        ws_extrato.column_dimensions['C'].width = 30
        ws_extrato.column_dimensions['D'].width = 18
        ws_extrato.column_dimensions['E'].width = 50

        conn.close()

        # Salva arquivo
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        # Gera nome do arquivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f'Faturamento_Rezende_{timestamp}.xlsx'

        # Envia arquivo
        return send_file(
            tmp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        return jsonify({'error': f'Erro ao exportar planilha completa: {str(e)}'}), 500


if __name__ == '__main__':
    # Cria pasta de uploads se não existir
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    app.run(debug=True, host='0.0.0.0', port=5000)