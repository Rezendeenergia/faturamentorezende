from openpyxl import load_workbook
from datetime import datetime


class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path, keep_vba=True)
        self.sheet_nfs = self.wb["NF'S"]

    def calcular_retencoes(self, tipo, valor_bruto):
        """Calcula todas as retenções baseado no tipo"""
        retencoes = {
            'inss': 0.0,
            'aliquota_inss': 0.0,
            'iss': 0.0,
            'aliquota_iss': 0.0,
            'retencao_equatorial': 0.0,
            'pis_cofins_csll': 0.0
        }

        if tipo == 'CONSTRUCAO':
            # INSS: 11% sobre 50% do valor (mão de obra)
            retencoes['inss'] = valor_bruto * 0.50 * 0.11
            retencoes['aliquota_inss'] = 0.055  # 5.5% sobre valor bruto
            # ISS: 5%
            retencoes['iss'] = valor_bruto * 0.05
            retencoes['aliquota_iss'] = 0.05
            # Retenção Equatorial: 5%
            retencoes['retencao_equatorial'] = valor_bruto * 0.05

        elif tipo == 'ENSAIO DIELETRICO':
            # INSS: 11% sobre valor total
            retencoes['inss'] = valor_bruto * 0.11
            retencoes['aliquota_inss'] = 0.11
            # ISS: 5%
            retencoes['iss'] = valor_bruto * 0.05
            retencoes['aliquota_iss'] = 0.05
            # Retenção Equatorial: 0%
            retencoes['retencao_equatorial'] = 0.0

        elif tipo == 'TRANSPORTE':
            # INSS: 0%
            retencoes['inss'] = 0.0
            retencoes['aliquota_inss'] = 0.0
            # ISS: 5%
            retencoes['iss'] = valor_bruto * 0.05
            retencoes['aliquota_iss'] = 0.05
            # Retenção Equatorial: 3%
            retencoes['retencao_equatorial'] = valor_bruto * 0.03

        elif tipo == 'TRANSPORTE_CTE':
            # INSS: 0%
            retencoes['inss'] = 0.0
            retencoes['aliquota_inss'] = 0.0
            # ISS: 0% (não tem ISS no CT-e)
            retencoes['iss'] = 0.0
            retencoes['aliquota_iss'] = 0.0
            # Retenção Equatorial: 3%
            retencoes['retencao_equatorial'] = valor_bruto * 0.03

        # PIS/COFINS/CSLL: 4.65% (apenas se marcado como retido)
        retencoes['pis_cofins_csll'] = valor_bruto * 0.0465

        return retencoes

    def calcular_valor_nominal(self, valor_bruto, retencoes, pis_cofins_retido=False):
        """Calcula o valor nominal (o que deveria estar no banco)"""
        valor_nominal = valor_bruto - retencoes['inss'] - retencoes['iss'] - retencoes['retencao_equatorial']

        if pis_cofins_retido:
            valor_nominal -= retencoes['pis_cofins_csll']

        return valor_nominal

    def inserir_nota(self, dados):
        """Insere uma nova nota fiscal na planilha"""
        # Encontra a próxima linha vazia
        next_row = self.sheet_nfs.max_row + 1

        # Calcula retenções
        retencoes = self.calcular_retencoes(dados['tipo'], dados['valor_bruto'])

        # Calcula valor nominal
        valor_nominal = self.calcular_valor_nominal(
            dados['valor_bruto'],
            retencoes,
            dados.get('pis_cofins_retido', False)
        )

        # Mapeia colunas (A=1, B=2, etc.)
        colunas = {
            'data_emissao': 1,  # A
            'numero_nf': 2,  # B
            'tipo': 3,  # C
            'valor_bruto': 4,  # D
            'localidade': 5,  # E
            'inss': 6,  # F
            'aliquota_inss': 7,  # G (fórmula)
            'iss': 8,  # H
            'aliquota_iss': 9,  # I (fórmula)
            'retencao_equatorial': 10,  # J (fórmula)
            'tomador': 11,  # K
            'pis_cofins_csll': 12,  # L
            'valor_nominal_conf': 13,  # M (manual)
            'valor_nominal': 14,  # N (calculado)
            'valor_liquido_vinci': 15,  # O
            'data_adiantamento': 16,  # P
            'percentual_adiantamento': 17,  # Q (fórmula)
            'valor_retido_vinci': 18  # R (fórmula)
        }

        # Insere dados
        self.sheet_nfs.cell(next_row, colunas['data_emissao']).value = datetime.strptime(dados['data_emissao'],
                                                                                         '%d/%m/%Y') if dados.get(
            'data_emissao') else None
        self.sheet_nfs.cell(next_row, colunas['numero_nf']).value = int(dados['numero_nf']) if dados.get(
            'numero_nf') else None
        self.sheet_nfs.cell(next_row, colunas['tipo']).value = dados['tipo']
        self.sheet_nfs.cell(next_row, colunas['valor_bruto']).value = dados['valor_bruto']
        self.sheet_nfs.cell(next_row, colunas['localidade']).value = dados.get('localidade', '')
        self.sheet_nfs.cell(next_row, colunas['inss']).value = retencoes['inss']
        self.sheet_nfs.cell(next_row, colunas['iss']).value = retencoes['iss']
        self.sheet_nfs.cell(next_row, colunas['tomador']).value = dados.get('tomador', '')

        # PIS/COFINS/CSLL apenas se retido
        if dados.get('pis_cofins_retido', False):
            self.sheet_nfs.cell(next_row, colunas['pis_cofins_csll']).value = retencoes['pis_cofins_csll']

        # Valor nominal calculado
        self.sheet_nfs.cell(next_row, colunas['valor_nominal']).value = valor_nominal

        # Valor nominal conferência (manual)
        if dados.get('valor_nominal_conferencia'):
            self.sheet_nfs.cell(next_row, colunas['valor_nominal_conf']).value = dados['valor_nominal_conferencia']

        # Dados de adiantamento
        if dados.get('foi_adiantado', False):
            if dados.get('data_adiantamento'):
                self.sheet_nfs.cell(next_row, colunas['data_adiantamento']).value = datetime.strptime(
                    dados['data_adiantamento'], '%d/%m/%Y')
            if dados.get('valor_liquido_vinci'):
                self.sheet_nfs.cell(next_row, colunas['valor_liquido_vinci']).value = dados['valor_liquido_vinci']

        # Fórmulas
        self._inserir_formulas(next_row, colunas, dados['tipo'])

        return next_row

    def _inserir_formulas(self, row, colunas, tipo):
        """Insere fórmulas nas colunas calculadas"""
        # Alíquota INSS
        self.sheet_nfs.cell(row, colunas['aliquota_inss']).value = f'=F{row}/D{row}'

        # Alíquota ISS
        self.sheet_nfs.cell(row, colunas['aliquota_iss']).value = f'=H{row}/D{row}'

        # Retenção Equatorial (baseado no tipo)
        if tipo == 'CONSTRUCAO':
            percentual = 0.05
        elif tipo in ['TRANSPORTE', 'TRANSPORTE_CTE']:
            percentual = 0.03
        else:
            percentual = 0.0

        self.sheet_nfs.cell(row, colunas['retencao_equatorial']).value = f'=D{row}*{percentual}'

        # Percentual de adiantamento
        self.sheet_nfs.cell(row, colunas['percentual_adiantamento']).value = f'=IF(O{row}>0,(N{row}-O{row})/N{row},"")'

        # Valor retido Vinci
        self.sheet_nfs.cell(row, colunas['valor_retido_vinci']).value = f'=IF(O{row}>0,N{row}-O{row},"")'

    def salvar(self):
        """Salva as alterações no Excel"""
        self.wb.save(self.excel_path)

    def fechar(self):
        """Fecha o workbook"""
        self.wb.close()